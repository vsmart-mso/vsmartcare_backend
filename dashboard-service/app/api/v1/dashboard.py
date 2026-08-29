"""สรุปจำนวนคำร้องสำหรับหน้า dashboard — รับ `province_id` ตรงจาก query param

ไม่มี permission/scope check ในตัว service นี้เอง (ตั้งใจ) — เหมือน pattern ของ
`case-service/app/api/v1/case_for_staff.py::list_cases_for_staff` ทุกประการ:
caller (BFF) ส่ง province_id/district_id/current_status_id ที่ "อนุญาตแล้ว" มาตรง ๆ
"""

from __future__ import annotations

import io
from datetime import date, datetime, timezone
from typing import Literal
from zoneinfo import ZoneInfo

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession

from ...core.database import get_session
from ...schemas import (
    DashboardCaseRow,
    DashboardCasesRead,
    DashboardDistrictRow,
    DashboardDistrictsRead,
    DashboardNationalOverviewRead,
    DashboardOverviewRead,
    DashboardProvinceRow,
    DashboardProvincesRead,
    DashboardStatusCount,
    DashboardSubDistrictRow,
    DashboardSubDistrictsRead,
)
from ...settings import settings
from ...queries import (
    DashboardCaseFilters,
    fetch_active_current_statuses,
    fetch_dashboard_case_export_rows,
    fetch_dashboard_case_list,
    fetch_dashboard_cases_count,
    fetch_dashboard_cases_page,
    fetch_dashboard_situation_export_rows,
    fetch_district,
    fetch_districts_page,
    fetch_districts_status_breakdown,
    fetch_districts_total_count,
    fetch_existing_province_ids,
    fetch_national_status_counts,
    fetch_national_total,
    fetch_overview_status_counts,
    fetch_overview_total,
    fetch_province,
    fetch_provinces_page,
    fetch_provinces_status_breakdown,
    fetch_provinces_total_count,
    fetch_sub_districts_page,
    fetch_sub_districts_status_breakdown,
    fetch_sub_districts_total_count,
)

router = APIRouter(prefix="/v1/dashboard", tags=["dashboard"])
BANGKOK_PROVINCE_ID = 10

BANGKOK_EXPORT_HEADERS = [
    "หมายเลขเคส",
    "ชื่อ",
    "นามสกุล",
    "หมวดเงิน",
    "หมวดเงิน_ชื่อเต็ม",
    "status_id",
    "สถานะ_ระบบ",
    "ขั้นตอน_ธุรกิจ",
    "ผู้รับผิดชอบ",
    "วันที่ยื่น",
]

PROVINCIAL_EXPORT_GROUPS = [
    ("A2:A3", "ลำดับ"),
    ("B2:B3", "ที่มา"),
    ("C2:C3", "เลขที่คำร้อง"),
    ("D2:D3", "วันที่แจ้งเรื่อง"),
    ("E2:G2", "รายละเอียดคำร้อง"),
    ("H2:U2", "ผู้ประสบปัญหาทางสังคมยื่นคำขอด้วยตนเอง"),
    ("V2:AA2", "ข้อมูลเศรษฐกิจครอบครัว"),
    ("AB2:AD2", "ข้อมูลสมาชิกและการอุปการะ"),
    ("AE2:AH2", "สิทธิสวัสดิการที่เคยได้รับ"),
    ("AI2:AM2", "สภาพปัญหาและความช่วยเหลือที่ต้องการ"),
    ("AN2:AQ2", "การพิจารณาให้ความช่วยเหลือ"),
    ("AR2:AY2", "ข้อมูลการจ่ายเงิน"),
    ("AZ2:BB2", "นักสังคมสงเคราะห์"),
]

PROVINCIAL_EXPORT_HEADERS = [
    "ลำดับ",
    "ที่มา",
    "เลขที่คำร้อง",
    "วันที่แจ้งเรื่อง",
    "กรณีฉุกเฉิน",
    "ประเภทเคส",
    "แหล่งข้อมูลเคสเดิม",
    "เลขประจำตัวประชาชน",
    "ชื่อ-นามสกุล",
    "วันเดือนปีเกิด",
    "อายุ",
    "เพศ",
    "ความสัมพันธ์",
    "สถานภาพ",
    "โทรศัพท์",
    "โทรสาร",
    "โทรศัพท์มือถือ",
    "อีเมล",
    "เป็นเจ้าหน้าที่รัฐ",
    "ที่อยู่ปัจจุบัน",
    "พิกัด",
    "อาชีพ",
    "อาชีพหลักของครอบครัว",
    "รายได้เฉลี่ย/เดือน",
    "ที่มาของรายได้",
    "สภาพที่อยู่อาศัย",
    "ค่าเช่า/เดือน",
    "การอุปการะ",
    "จำนวนสมาชิกครอบครัว",
    "รายละเอียดสมาชิกครอบครัว",
    "สถานะการรับสวัสดิการ",
    "จำนวนครั้ง",
    "จำนวนเงินรวม",
    "ประเภทสวัสดิการ",
    "สภาพปัญหาความเดือดร้อน",
    "ความช่วยเหลือที่ต้องการ",
    "จำนวนเงิน",
    "รายละเอียดสิ่งของ",
    "รายละเอียดความช่วยเหลืออื่น",
    "ความเห็นวินิจฉัย",
    "จำนวนเงินที่วินิจฉัย",
    "ระเบียบ/หลักเกณฑ์",
    "ประเภทเงิน",
    "วิธีการจ่ายเงิน",
    "ชื่อผู้รับเงิน",
    "เลขบัตรผู้รับเงิน",
    "โทรศัพท์ผู้รับเงิน",
    "ธนาคาร",
    "เลขที่บัญชี",
    "ชื่อบัญชี",
    "สาขา",
    "นักสังคมสงเคราะห์",
    "ตำแหน่ง",
    "เลขใบอนุญาต",
]


def _clean_ids(ids: list[int] | None) -> list[int] | None:
    """list ว่าง (`?type_money_id=` ไม่ระบุค่า) ให้ถือว่าไม่ได้กรอง เหมือน None."""
    return ids or None


def _clean_text_filter(value: str | None) -> str | None:
    if value is None:
        return None
    value = value.strip()
    return value or None


_BANGKOK = ZoneInfo("Asia/Bangkok")


def _sla_fields(
    started_at: datetime | None,
    sla_days: int | None,
    *,
    completed_at: datetime | None,
    frozen_elapsed: int | None,
) -> dict:
    """คำนวณฟิลด์ SLA ให้สอดคล้อง case-service `process_sla_fields_dict`."""
    empty = {
        "process_started_at": started_at,
        "process_completed_at": completed_at,
        "process_sla_days": sla_days,
        "process_elapsed_days": None,
        "process_remaining_days": None,
        "process_traffic_color": None,
        "process_is_overdue": None,
        "time_count_process": None,
    }
    if sla_days is None:
        return empty

    def _now_bangkok(value: datetime | None) -> datetime:
        if value is None:
            return datetime.now(tz=_BANGKOK)
        if value.tzinfo is None:
            return value.replace(tzinfo=_BANGKOK)
        return value.astimezone(_BANGKOK)

    def _elapsed(start: datetime, end: datetime | None) -> int:
        end_bkk = _now_bangkok(end)
        start_bkk = start.astimezone(_BANGKOK) if start.tzinfo else start.replace(tzinfo=_BANGKOK)
        return (end_bkk.date() - start_bkk.date()).days

    def _traffic(elapsed: int, sla: int) -> str:
        if sla == 10:
            if elapsed <= 4:
                return "green"
            if elapsed <= 7:
                return "yellow"
            if elapsed <= 10:
                return "orange"
            return "red"
        if sla == 15:
            if elapsed <= 5:
                return "green"
            if elapsed <= 11:
                return "yellow"
            if elapsed <= 15:
                return "orange"
            return "red"
        if sla <= 0:
            return "red"
        pct = elapsed / sla
        if pct <= 0.40:
            return "green"
        if pct <= 0.70:
            return "yellow"
        if pct <= 1.0:
            return "orange"
        return "red"

    if started_at is None:
        if completed_at is not None and frozen_elapsed is not None:
            remaining = sla_days - frozen_elapsed
            return {
                "process_started_at": started_at,
                "process_completed_at": completed_at,
                "process_sla_days": sla_days,
                "process_elapsed_days": frozen_elapsed,
                "process_remaining_days": remaining,
                "process_traffic_color": _traffic(frozen_elapsed, sla_days),
                "process_is_overdue": remaining < 0,
                "time_count_process": frozen_elapsed,
            }
        return empty

    elapsed = _elapsed(started_at, completed_at)
    remaining = sla_days - elapsed
    return {
        "process_started_at": started_at,
        "process_completed_at": completed_at,
        "process_sla_days": sla_days,
        "process_elapsed_days": elapsed,
        "process_remaining_days": remaining,
        "process_traffic_color": _traffic(elapsed, sla_days),
        "process_is_overdue": remaining < 0,
        "time_count_process": elapsed,
    }


def _row_to_dashboard_case(row: dict) -> DashboardCaseRow:
    completed_at = row.get("process_completed_at")
    frozen = row.get("time_count_process") if completed_at is not None else None
    row.update(
        _sla_fields(
            row.get("process_started_at"),
            row.get("process_sla_days"),
            completed_at=completed_at,
            frozen_elapsed=frozen if isinstance(frozen, int) else None,
        )
    )
    row["current_address_province_id"] = row.get("province_id")
    row["current_address_province_name"] = row.get("province_name")
    row["person_age"] = int(row.get("person_age") or 0)
    row["count_037"] = int(row.get("count_037") or 0)
    row["count_038"] = int(row.get("count_038") or 0)
    row.setdefault("prior_self_submit_case_numbers", [])
    row.setdefault("self_submit_fiscal_year_count", 0)
    row.setdefault("self_submit_fiscal_year_case_numbers", [])
    row.setdefault("responsible_division_name", None)

    sources = row.get("existing_case_detected_sources")
    if sources is not None and not isinstance(sources, list):
        row["existing_case_detected_sources"] = None

    require_ktb = bool(row.get("require_ktb_corporate", True))
    source = row.get("existing_case_source")
    ref_id = row.get("existing_case_ref_id")
    prior_reuse = None
    if not require_ktb and source in {"VCARE", "Legacy"} and ref_id is not None:
        prior_reuse = int(ref_id)
    row["prior_ktb_reuse_applicant_id"] = prior_reuse
    return DashboardCaseRow.model_validate(row)


async def _require_province(session: AsyncSession, province_id: int) -> dict:
    province = await fetch_province(session, province_id)
    if province is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="province_not_found")
    return province


# 77 = จำนวนจังหวัดทั้งประเทศ — ส่งเกินนี้แปลว่า caller ส่งอะไรผิดแน่ ๆ
MAX_PROVINCE_FILTER = 77


async def _validate_province_ids(
    session: AsyncSession, province_id: list[int] | None
) -> list[int] | None:
    """ตรวจ `province_id` หลายค่าก่อนใช้กรอง — ห้ามเมิน param แล้วคืนข้อมูลทั้งประเทศ

    ไม่ส่งมา (หรือส่งมาว่าง) = ไม่กรอง คืน None ตามพฤติกรรมเดิม
    ส่งมาแล้วมีค่าผิด ต้อง error พร้อมบอกว่าค่าไหนผิด เพื่อให้ caller debug ได้
    """
    ids = _clean_ids(province_id)
    if ids is None:
        return None

    unique_ids = sorted(set(ids))
    if len(unique_ids) > MAX_PROVINCE_FILTER:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=(
                f"too_many_province_id: ส่งมา {len(unique_ids)} ค่า "
                f"เกินที่รองรับ {MAX_PROVINCE_FILTER} ค่า"
            ),
        )

    existing = await fetch_existing_province_ids(session, unique_ids)
    missing = [pid for pid in unique_ids if pid not in existing]
    if missing:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"province_not_found: {missing}",
        )
    return unique_ids


def _fmt_date(value) -> str:
    if value is None:
        return ""
    try:
        return value.strftime("%d/%m/%Y")
    except AttributeError:
        return str(value)


def _fmt_money(value) -> str:
    if value is None:
        return ""
    return float(value)


def _yes_no(value) -> str:
    return "ใช่" if value else "ไม่ใช่"


def _existing_case_label(value) -> str:
    if value is True:
        return "รายเดิม"
    if value is False:
        return "รายใหม่"
    return "ไม่มีข้อมูล"


def _join_address(row: dict) -> str:
    parts = [
        row.get("house_number"),
        f"หมู่ {row.get('house_moo')}" if row.get("house_moo") else None,
        row.get("house_name"),
        f"ตรอก{row.get('alley')}" if row.get("alley") else None,
        f"ซอย{row.get('sub_lane')}" if row.get("sub_lane") else None,
        f"ถนน{row.get('road')}" if row.get("road") else None,
        f"ต.{row.get('sub_district_name')}" if row.get("sub_district_name") else None,
        f"อ.{row.get('district_name')}" if row.get("district_name") else None,
        f"จ.{row.get('province_name')}" if row.get("province_name") else None,
        row.get("postcode"),
        row.get("nearby_landmark"),
    ]
    return " ".join(str(part) for part in parts if part)


def _coordinate(row: dict) -> str:
    lat = row.get("latitude")
    lng = row.get("longitude")
    return f"{lat}, {lng}" if lat and lng else ""


def _existing_case_source(row: dict) -> str:
    source_labels = {
        "vcare_main": "พม.CARE",
        "vsmart_main": "vSmart",
        "mso_logbook": "สปสช. (Welfare)",
    }
    sources = row.get("existing_case_detected_sources")
    if isinstance(sources, list) and sources:
        return ", ".join(source_labels.get(str(item), str(item)) for item in sources)
    source = row.get("existing_case_source")
    return source_labels.get(str(source), str(source)) if source else "ไม่มีข้อมูล"


def _responsible_person(row: dict) -> str:
    return (
        row.get("responsible_person")
        or row.get("latest_status_update_by")
        or row.get("sw_user_sdshv")
        or row.get("sw_explorer_sdshv")
        or row.get("social_worker_name")
        or ""
    )


def _make_bangkok_workbook(rows: list[dict]):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "ตัวอย่าง"
    ws.append(BANGKOK_EXPORT_HEADERS)
    for row in rows:
        ws.append(
            [
                row.get("case_number") or "",
                row.get("first_name") or "",
                row.get("last_name") or "",
                row.get("type_money_category_acronym") or "",
                row.get("type_money_category_name") or "",
                row.get("current_status_id") or "",
                row.get("current_status_label") or "",
                row.get("current_status_business_step")
                or row.get("current_status_public_label")
                or "",
                _responsible_person(row),
                _fmt_date(row.get("created_at")),
            ]
        )

    header_fill = PatternFill("solid", fgColor="1F4E78")
    alternate_fill = PatternFill("solid", fgColor="F4F8FB")
    thin = Side(style="thin", color="D9E2F3")
    border = Border(bottom=thin)
    for cell in ws[1]:
        cell.font = Font(name="TH SarabunPSK", size=16, bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        for cell in row:
            cell.font = Font(name="TH SarabunPSK", size=16)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            if row_idx % 2 == 0:
                cell.fill = alternate_fill
    for col_idx, col in enumerate(ws.columns, start=1):
        width = min(max(len(str(cell.value or "")) for cell in col) + 2, 35)
        ws.column_dimensions[col[0].column_letter].width = max(width, 10)
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 90
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = "1:1"
    return wb


def _provincial_export_title(
    rows: list[dict],
    province_ids: list[int] | None,
    province_names: list[str] | None,
) -> str:
    names = province_names or sorted({row.get("province_name") for row in rows if row.get("province_name")})
    if len(names) == 1:
        return f"สำนักงานพัฒนาสังคมและความมั่นคงของมนุษย์จังหวัด{names[0]}"
    if province_ids:
        return "รายงานข้อมูลคำร้อง พม.CARE"
    return "รายงานข้อมูลคำร้อง พม.CARE ทุกจังหวัด"


def _provincial_row_values(index: int, row: dict) -> list:
    full_name = " ".join(part for part in [row.get("first_name"), row.get("last_name")] if part)
    welfare_status = "เคยได้รับ" if row.get("has_received_welfare") else "ไม่เคยได้รับ"
    distress = row.get("family_distress") or row.get("problem_details") or ""
    phone = row.get("home_phone") or row.get("address_mobile_phone") or ""
    mobile = row.get("mobile_phone") or row.get("address_mobile_phone") or ""
    return [
        index,
        "พม.CARE",
        row.get("case_number") or "",
        _fmt_date(row.get("created_at")),
        _yes_no(row.get("is_emergency")),
        _existing_case_label(row.get("is_existing_case")),
        _existing_case_source(row),
        row.get("cid") or "",
        full_name,
        _fmt_date(row.get("birth_date")),
        row.get("applicant_age") or "",
        row.get("gender") or "",
        row.get("requester_relation_name") or "",
        row.get("marital_status_name") or "",
        phone,
        row.get("fax_number") or "",
        mobile,
        row.get("email_address") or "",
        _yes_no(row.get("is_government_officer")),
        _join_address(row),
        _coordinate(row),
        row.get("occupation_name") or "",
        row.get("family_occupation_name") or "",
        _fmt_money(row.get("monthly_income")),
        row.get("income_source_names") or "",
        row.get("housing_shelter") or row.get("housing_type_name") or "",
        _fmt_money(row.get("housing_types_rent")),
        row.get("dependency_names") or "",
        row.get("household_members_count") or "",
        row.get("household_member_details") or "",
        welfare_status,
        row.get("received_count") or "",
        _fmt_money(row.get("total_received_amount")),
        row.get("received_welfare_type_names") or "",
        distress,
        row.get("request_type_names") or "",
        "",
        row.get("request_in_kind_text") or "",
        row.get("request_other_text") or "",
        row.get("diagnosis_text") or row.get("regulation_comment") or "",
        _fmt_money(row.get("approved_money_amount")),
        row.get("regulation_name") or "",
        row.get("type_money_name") or row.get("type_money_category_name") or "",
        row.get("payment_method_name") or "",
        row.get("payee_name") or full_name,
        row.get("payee_cid") or row.get("cid") or "",
        mobile,
        row.get("payment_bank_name") or "",
        row.get("account_number") or "",
        row.get("account_name") or "",
        row.get("bank_branch") or "",
        row.get("social_worker_name") or row.get("sw_user_sdshv") or row.get("sw_explorer_sdshv") or "",
        row.get("social_worker_position") or "",
        "",
    ]


def _make_provincial_workbook(
    rows: list[dict],
    province_ids: list[int] | None,
    province_names: list[str] | None = None,
):
    import xlsxwriter

    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output, {"in_memory": True})
    worksheet = workbook.add_worksheet()
    merge_format = workbook.add_format(
        {
            "bold": 1,
            "border": 1,
            "align": "center",
            "valign": "vcenter",
            "font_name": "TH SarabunPSK",
            "font_size": 18,
        }
    )
    text_cell_format = workbook.add_format(
        {
            "border": 1,
            "valign": "vcenter",
            "num_format": "@",
            "font_name": "TH SarabunPSK",
            "font_size": 16,
        }
    )

    worksheet.merge_range(
        "A1:BB1",
        _provincial_export_title(rows, province_ids, province_names),
        merge_format,
    )
    for merged_range, label in PROVINCIAL_EXPORT_GROUPS:
        worksheet.merge_range(merged_range, label, merge_format)
    for col_idx, label in enumerate(PROVINCIAL_EXPORT_HEADERS):
        if col_idx >= 4:
            worksheet.write(2, col_idx, label, merge_format)

    for row_idx, row in enumerate(rows, start=3):
        values = _provincial_row_values(row_idx - 2, row)
        for col_idx, value in enumerate(values):
            worksheet.write_string(row_idx, col_idx, "" if value is None else str(value), text_cell_format)

    worksheet.autofit()
    worksheet.set_column("A:A", 6)
    workbook.close()
    output.seek(0)
    return output


def _workbook_response(wb, *, safe_filename: str, utf8_filename: str) -> StreamingResponse:
    from urllib.parse import quote

    if isinstance(wb, io.BytesIO):
        buf = wb
    else:
        buf = io.BytesIO()
        wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": (
                f'attachment; filename="{safe_filename}"; filename*=UTF-8\'\'{quote(utf8_filename)}'
            )
        },
    )


@router.get("/national/overview", response_model=DashboardNationalOverviewRead)
async def get_national_overview(
    province_id: list[int] | None = Query(
        None, description="กรองเฉพาะจังหวัดที่ระบุ ส่งซ้ำได้หลายค่า ไม่ส่ง = ทุกจังหวัด"
    ),
    current_status_id: list[int] | None = Query(
        None, description="กรองตาม current_status_id ได้หลายค่า"
    ),
    type_money_id: list[int] | None = Query(None),
    session: AsyncSession = Depends(get_session),
) -> DashboardNationalOverviewRead:
    """สรุปทั้งประเทศ (หรือเฉพาะจังหวัดที่ระบุ) แยกตาม status — ใช้ทำ donut chart."""
    province_ids = await _validate_province_ids(session, province_id)
    current_status_ids = _clean_ids(current_status_id)
    type_money_ids = _clean_ids(type_money_id)
    total = await fetch_national_total(
        session,
        province_ids=province_ids,
        current_status_ids=current_status_ids,
        type_money_ids=type_money_ids,
    )
    status_rows = await fetch_national_status_counts(
        session,
        province_ids=province_ids,
        current_status_ids=current_status_ids,
        type_money_ids=type_money_ids,
    )
    statuses = [
        DashboardStatusCount(
            current_status_id=row["current_status_id"],
            label=row["label"],
            color=row["color"],
            count=row["count"],
            percent=round((row["count"] / total * 100), 1) if total else 0.0,
        )
        for row in status_rows
    ]
    return DashboardNationalOverviewRead(
        total=total,
        updated_at=datetime.now(timezone.utc),
        statuses=statuses,
    )


async def _build_provinces_response(
    session: AsyncSession,
    *,
    province_ids: list[int] | None,
    current_status_id: list[int] | None,
    type_money_id: list[int] | None,
    page: int,
    page_size: int,
    _prefetched_total: int | None = None,
) -> DashboardProvincesRead:
    current_status_ids = _clean_ids(current_status_id)
    type_money_ids = _clean_ids(type_money_id)

    # total_items/total_pages ต้องนับจากชุดที่กรองแล้ว ไม่ใช่ 77 จังหวัดเสมอ
    total_items = (
        _prefetched_total
        if _prefetched_total is not None
        else await fetch_provinces_total_count(session, province_ids=province_ids)
    )
    total_pages = max(1, (total_items + page_size - 1) // page_size) if total_items else 1

    province_rows = await fetch_provinces_page(
        session,
        province_ids=province_ids,
        current_status_ids=current_status_ids,
        type_money_ids=type_money_ids,
        limit=page_size,
        offset=(page - 1) * page_size,
    )
    page_province_ids = [r["province_id"] for r in province_rows]

    breakdown_rows = await fetch_provinces_status_breakdown(
        session,
        province_ids=page_province_ids,
        current_status_ids=current_status_ids,
        type_money_ids=type_money_ids,
    )
    status_counts_by_province: dict[int, dict[str, int]] = {}
    for row in breakdown_rows:
        status_counts_by_province.setdefault(row["province_id"], {})[
            str(row["current_status_id"])
        ] = row["count"]

    items = [
        DashboardProvinceRow(
            province_id=row["province_id"],
            province_name=row["province_name"],
            status_counts=status_counts_by_province.get(row["province_id"], {}),
            total=row["total"],
        )
        for row in province_rows
    ]
    return DashboardProvincesRead(
        page=page,
        page_size=page_size,
        total_items=total_items,
        total_pages=total_pages,
        items=items,
    )


@router.get("/provinces", response_model=DashboardProvincesRead)
async def get_provinces(
    province_id: list[int] | None = Query(
        None, description="กรองเฉพาะจังหวัดที่ระบุ ส่งซ้ำได้หลายค่า ไม่ส่ง = ทุกจังหวัด"
    ),
    current_status_id: list[int] | None = Query(None),
    type_money_id: list[int] | None = Query(None),
    page: int = Query(1, ge=1),
    page_size: int | None = Query(None, ge=1),
    session: AsyncSession = Depends(get_session),
) -> DashboardProvincesRead:
    """ตารางรายจังหวัด (ทั้งประเทศ หรือเฉพาะที่ระบุ) แยกตาม status — มี pagination."""
    resolved = min(page_size or settings.default_page_size, settings.max_page_size)
    province_ids = await _validate_province_ids(session, province_id)
    return await _build_provinces_response(
        session,
        province_ids=province_ids,
        current_status_id=current_status_id,
        type_money_id=type_money_id,
        page=page,
        page_size=resolved,
    )


@router.get("/provinces/export")
async def export_provinces(
    province_id: list[int] | None = Query(
        None, description="กรองเฉพาะจังหวัดที่ระบุ ส่งซ้ำได้หลายค่า ไม่ส่ง = ทุกจังหวัด"
    ),
    current_status_id: list[int] | None = Query(None),
    type_money_id: list[int] | None = Query(None),
    session: AsyncSession = Depends(get_session),
) -> StreamingResponse:
    """Excel รายจังหวัด (ทั้งประเทศ หรือเฉพาะที่ระบุ) — ไม่มี pagination."""
    from openpyxl import Workbook
    from urllib.parse import quote

    province_ids = await _validate_province_ids(session, province_id)
    total = await fetch_provinces_total_count(session, province_ids=province_ids)
    data = await _build_provinces_response(
        session,
        province_ids=province_ids,
        current_status_id=current_status_id,
        type_money_id=type_money_id,
        page=1,
        page_size=max(total, 1),
        _prefetched_total=total,
    )
    status_labels = await fetch_active_current_statuses(session)
    status_ids = [r["id"] for r in status_labels]
    status_headers = [r["label"] for r in status_labels]

    wb = Workbook()
    ws = wb.active
    ws.title = "provinces"
    ws.append(["ลำดับ", "จังหวัด", *status_headers, "รวม"])
    for idx, row in enumerate(data.items, start=1):
        ws.append([idx, row.province_name, *[row.status_counts.get(str(s), 0) for s in status_ids], row.total])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    # กรองจังหวัดแล้วอย่าตั้งชื่อไฟล์ว่า "ทั้งประเทศ" — ผู้ใช้ สสว. จะเข้าใจผิดว่าได้ข้อมูลครบ
    scope = "ทั้งประเทศ" if province_ids is None else f"{len(province_ids)}-จังหวัด"
    safe = "dashboard-national-provinces.xlsx" if province_ids is None else "dashboard-provinces.xlsx"
    utf8 = quote(f"dashboard-{scope}-รายจังหวัด.xlsx")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{safe}"; filename*=UTF-8\'\'{utf8}'},
    )


@router.get("/cases")
async def get_dashboard_cases(
    province_id: list[int] | None = Query(None),
    district_id: int | None = Query(None, ge=1),
    current_status_id: list[int] | None = Query(None),
    type_money_id: list[int] | None = Query(None),
    search: str | None = Query(None, max_length=100),
    date_from: date | None = Query(None),
    date_to: date | None = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(10, ge=1, le=50),
    session: AsyncSession = Depends(get_session),
) -> dict:
    """รายการเคสรายบุคคลระดับประเทศ จังหวัด หรืออำเภอสำหรับ Modal แผนที่."""
    province_ids = _clean_ids(province_id)
    for selected_province_id in province_ids or []:
        await _require_province(session, selected_province_id)
    if date_from and date_to and date_from > date_to:
        raise HTTPException(status_code=400, detail="date_from ต้องไม่มากกว่า date_to")
    result = await fetch_dashboard_case_list(
        session,
        province_ids=province_ids,
        district_id=district_id,
        current_status_ids=_clean_ids(current_status_id),
        type_money_ids=_clean_ids(type_money_id),
        search=(search or "").strip() or None,
        date_from=date_from,
        date_to=date_to,
        limit=page_size,
        offset=(page - 1) * page_size,
    )
    total_items = result["total_items"]
    return {
        "page": page,
        "page_size": page_size,
        "total_items": total_items,
        "total_pages": max(1, (total_items + page_size - 1) // page_size),
        "items": result["items"],
    }


@router.get("/status/export")
async def export_status(
    level: Literal["provinces", "districts"] = Query(
        ..., description="provinces = Excel รายคำร้องระดับประเทศ, districts = Excel รายคำร้องในจังหวัด"
    ),
    report_type: Literal["situation", "case"] = Query(
        "case", description="situation = รูปแบบ 10 คอลัมน์, case = รูปแบบรายงานเคส 54 คอลัมน์"
    ),
    province_id: list[int] | None = Query(
        None, description="provinces ส่งซ้ำได้หลายจังหวัด; districts ต้องส่ง 1 จังหวัด"
    ),
    applicant_id: list[int] | None = Query(
        None, description="กรองเฉพาะ applicant ที่เลือกจากหน้าเอกสารประชุม"
    ),
    district_id: list[int] | None = Query(None),
    sub_district_id: list[int] | None = Query(None),
    current_status_id: list[int] | None = Query(None),
    type_money_id: list[int] | None = Query(None),
    session: AsyncSession = Depends(get_session),
) -> StreamingResponse:
    """Excel จาก dashboard — กรุงเทพฯ ใช้ layout 10 คอลัมน์, จังหวัดอื่นใช้ layout 76 จังหวัด."""
    province_ids: list[int] | None
    if level == "provinces":
        province_ids = await _validate_province_ids(session, province_id)
    else:
        ids = _clean_ids(province_id)
        if not ids or len(ids) != 1:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="districts_export_requires_one_province_id",
            )
        province_ids = await _validate_province_ids(session, ids)

    current_status_ids = _clean_ids(current_status_id)
    type_money_ids = _clean_ids(type_money_id)
    applicant_ids = _clean_ids(applicant_id)
    district_ids = _clean_ids(district_id)
    sub_district_ids = _clean_ids(sub_district_id)
    if report_type == "situation":
        rows = await fetch_dashboard_situation_export_rows(
            session,
            province_ids=province_ids,
            applicant_ids=applicant_ids,
            district_ids=district_ids,
            sub_district_ids=sub_district_ids,
            current_status_ids=current_status_ids,
            type_money_ids=type_money_ids,
        )
    else:
        rows = await fetch_dashboard_case_export_rows(
            session,
            province_ids=province_ids,
            district_ids=district_ids,
            sub_district_ids=sub_district_ids,
            current_status_ids=current_status_ids,
            type_money_ids=type_money_ids,
            exclude_province_ids=None,
        )
    if not rows:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="dashboard_export_no_data",
        )
    if report_type == "situation":
        wb = _make_bangkok_workbook(rows)
        return _workbook_response(
            wb,
            safe_filename="dashboard-situation-report.xlsx",
            utf8_filename="dashboard-รายงานสถานการณ์.xlsx",
        )

    province_names = None
    if province_ids:
        province_names = []
        for province_id in province_ids:
            province = await fetch_province(session, province_id)
            if province and province.get("name"):
                province_names.append(province["name"])

    wb = _make_provincial_workbook(rows, province_ids, province_names)
    safe_scope = "selected-area" if province_ids else "all-provinces"
    thai_scope = "พื้นที่ที่เลือก" if province_ids else "ทุกจังหวัด"
    return _workbook_response(
        wb,
        safe_filename=f"dashboard-{safe_scope}-cases.xlsx",
        utf8_filename=f"dashboard-{thai_scope}-รายคำร้อง.xlsx",
    )


@router.get("/overview", response_model=DashboardOverviewRead)
async def get_overview(
    province_id: int = Query(..., description="รหัสจังหวัดที่ต้องการดู"),
    type_money_id: list[int] | None = Query(
        None, description="กรองตาม type_money_category.id ได้หลายค่า"
    ),
    session: AsyncSession = Depends(get_session),
) -> DashboardOverviewRead:
    province = await _require_province(session, province_id)
    type_money_ids = _clean_ids(type_money_id)

    total = await fetch_overview_total(
        session, province_id=province_id, type_money_ids=type_money_ids
    )
    status_rows = await fetch_overview_status_counts(
        session, province_id=province_id, type_money_ids=type_money_ids
    )

    statuses = [
        DashboardStatusCount(
            current_status_id=row["current_status_id"],
            label=row["label"],
            color=row["color"],
            count=row["count"],
            percent=round((row["count"] / total * 100), 1) if total else 0.0,
        )
        for row in status_rows
    ]

    return DashboardOverviewRead(
        province_id=province["id"],
        province_name=province["name"],
        total=total,
        updated_at=datetime.now(timezone.utc),
        statuses=statuses,
    )


async def _build_districts_response(
    session: AsyncSession,
    *,
    province_id: int,
    current_status_id: list[int] | None,
    type_money_id: list[int] | None,
    page: int,
    page_size: int,
    _prefetched_total: int | None = None,
) -> DashboardDistrictsRead:
    province = await _require_province(session, province_id)
    current_status_ids = _clean_ids(current_status_id)
    type_money_ids = _clean_ids(type_money_id)

    total_items = _prefetched_total if _prefetched_total is not None else await fetch_districts_total_count(session, province_id=province_id)
    total_pages = max(1, (total_items + page_size - 1) // page_size) if total_items else 1

    district_rows = await fetch_districts_page(
        session,
        province_id=province_id,
        current_status_ids=current_status_ids,
        type_money_ids=type_money_ids,
        limit=page_size,
        offset=(page - 1) * page_size,
    )
    district_ids = [r["district_id"] for r in district_rows]

    breakdown_rows = await fetch_districts_status_breakdown(
        session,
        province_id=province_id,
        district_ids=district_ids,
        current_status_ids=current_status_ids,
        type_money_ids=type_money_ids,
    )
    status_counts_by_district: dict[int, dict[str, int]] = {}
    for row in breakdown_rows:
        status_counts_by_district.setdefault(row["district_id"], {})[
            str(row["current_status_id"])
        ] = row["count"]

    items = [
        DashboardDistrictRow(
            district_id=row["district_id"],
            district_code=row.get("district_code"),
            district_name=row["district_name"],
            status_counts=status_counts_by_district.get(row["district_id"], {}),
            total=row["total"],
        )
        for row in district_rows
    ]

    return DashboardDistrictsRead(
        province_id=province["id"],
        province_name=province["name"],
        page=page,
        page_size=page_size,
        total_items=total_items,
        total_pages=total_pages,
        items=items,
    )


@router.get("/districts", response_model=DashboardDistrictsRead)
async def get_districts(
    province_id: int = Query(..., description="รหัสจังหวัดที่ต้องการดู"),
    current_status_id: list[int] | None = Query(
        None, description="กรองตาม current_status_id ได้หลายค่า"
    ),
    type_money_id: list[int] | None = Query(
        None, description="กรองตาม type_money_category.id ได้หลายค่า"
    ),
    page: int = Query(1, ge=1),
    page_size: int | None = Query(None, ge=1, description="ค่าเริ่มต้น/สูงสุดตั้งค่าผ่าน env"),
    session: AsyncSession = Depends(get_session),
) -> DashboardDistrictsRead:
    resolved_page_size = min(page_size or settings.default_page_size, settings.max_page_size)
    return await _build_districts_response(
        session,
        province_id=province_id,
        current_status_id=current_status_id,
        type_money_id=type_money_id,
        page=page,
        page_size=resolved_page_size,
    )


@router.get("/sub-districts", response_model=DashboardSubDistrictsRead)
async def get_sub_districts(
    district_id: int = Query(..., description="รหัสอำเภอที่ต้องการดู"),
    province_id: int = Query(..., description="รหัสจังหวัด (ตรวจสอบว่าอำเภออยู่ในจังหวัดนี้)"),
    current_status_id: list[int] | None = Query(None),
    type_money_id: list[int] | None = Query(None),
    page: int = Query(1, ge=1),
    page_size: int | None = Query(None, ge=1),
    session: AsyncSession = Depends(get_session),
) -> DashboardSubDistrictsRead:
    """ตารางรายตำบล แยกตาม status — คืนทุกตำบลในอำเภอ (แม้ count=0) พร้อม status_counts."""
    district = await fetch_district(session, district_id=district_id, province_id=province_id)
    if district is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="district_not_found")

    resolved_page_size = min(page_size or settings.default_page_size, settings.max_page_size)
    current_status_ids = _clean_ids(current_status_id)
    type_money_ids = _clean_ids(type_money_id)

    total_items = await fetch_sub_districts_total_count(session, district_id=district_id)
    total_pages = max(1, (total_items + resolved_page_size - 1) // resolved_page_size) if total_items else 1

    sub_district_rows = await fetch_sub_districts_page(
        session,
        district_id=district_id,
        province_id=province_id,
        current_status_ids=current_status_ids,
        type_money_ids=type_money_ids,
        limit=resolved_page_size,
        offset=(page - 1) * resolved_page_size,
    )
    sub_district_ids = [r["sub_district_id"] for r in sub_district_rows]

    breakdown_rows = await fetch_sub_districts_status_breakdown(
        session,
        district_id=district_id,
        province_id=province_id,
        sub_district_ids=sub_district_ids,
        current_status_ids=current_status_ids,
        type_money_ids=type_money_ids,
    )
    status_counts_by_sub_district: dict[int, dict[str, int]] = {}
    for row in breakdown_rows:
        status_counts_by_sub_district.setdefault(row["sub_district_id"], {})[
            str(row["current_status_id"])
        ] = row["count"]

    items = [
        DashboardSubDistrictRow(
            sub_district_id=row["sub_district_id"],
            sub_district_code=row.get("sub_district_code"),
            sub_district_name=row["sub_district_name"],
            status_counts=status_counts_by_sub_district.get(row["sub_district_id"], {}),
            total=row["total"],
        )
        for row in sub_district_rows
    ]

    return DashboardSubDistrictsRead(
        district_id=district["district_id"],
        district_name=district["district_name"],
        province_id=district["province_id"],
        province_name=district["province_name"],
        page=page,
        page_size=resolved_page_size,
        total_items=total_items,
        total_pages=total_pages,
        items=items,
    )


@router.get("/districts/export")
async def export_districts(
    province_id: int = Query(..., description="รหัสจังหวัดที่ต้องการดู"),
    current_status_id: list[int] | None = Query(None),
    type_money_id: list[int] | None = Query(None),
    session: AsyncSession = Depends(get_session),
) -> StreamingResponse:
    """Excel รายอำเภอ — filter เดียวกับ `/districts` แต่ดึงทุกอำเภอในจังหวัด (ไม่ pagination)."""
    from openpyxl import Workbook

    total_districts = await fetch_districts_total_count(session, province_id=province_id)
    data = await _build_districts_response(
        session,
        province_id=province_id,
        current_status_id=current_status_id,
        type_money_id=type_money_id,
        page=1,
        page_size=max(total_districts, 1),
        _prefetched_total=total_districts,
    )

    status_labels = await fetch_active_current_statuses(session)
    status_ids = [row["id"] for row in status_labels]
    status_headers = [row["label"] for row in status_labels]

    wb = Workbook()
    ws = wb.active
    ws.title = "districts"
    ws.append(["ลำดับ", "อำเภอ", *status_headers, "รวม"])
    for idx, row in enumerate(data.items, start=1):
        ws.append(
            [
                idx,
                row.district_name,
                *[row.status_counts.get(str(sid), 0) for sid in status_ids],
                row.total,
            ]
        )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    # ชื่อจังหวัดเป็นภาษาไทย — header ASCII-only ใช้ filename สำรอง, ตัวจริง (UTF-8) ใช้ filename*
    # ตาม RFC 6266 เพื่อให้เบราว์เซอร์ถอดชื่อไฟล์ภาษาไทยได้ถูกต้อง
    from urllib.parse import quote

    safe_filename = f"dashboard-province-{province_id}-districts.xlsx"
    utf8_filename = quote(f"dashboard-{data.province_name}-districts.xlsx")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": (
                f'attachment; filename="{safe_filename}"; filename*=UTF-8\'\'{utf8_filename}'
            )
        },
    )


@router.get("/cases", response_model=DashboardCasesRead)
async def get_cases(
    province_id: list[int] | None = Query(
        None, description="กรองเฉพาะจังหวัดที่ระบุ ส่งซ้ำได้หลายค่า ไม่ส่ง = ทุกจังหวัด"
    ),
    current_status_id: list[int] | None = Query(
        None, description="กรองตาม current_status_id ได้หลายค่า"
    ),
    type_money_id: list[int] | None = Query(
        None, description="กรองตาม type_money_category.id ได้หลายค่า"
    ),
    case_number: str | None = Query(None, description="ค้นหาจากเลข case"),
    current_status: str | None = Query(None, description="ค้นหาจากข้อความสถานะฝั่งเจ้าหน้าที่"),
    firstname: str | None = Query(None, description="ค้นหาจากชื่อ"),
    lastname: str | None = Query(None, description="ค้นหาจากนามสกุล"),
    cid: str | None = Query(None, description="ค้นหาจากเลขบัตรประชาชน"),
    datetime_create: date | None = Query(None, description="วันที่สร้าง case (YYYY-MM-DD)"),
    province_name: str | None = Query(None, description="ค้นหาจากชื่อจังหวัด"),
    district_id: int | None = Query(None, description="กรองตามอำเภอ"),
    district_name: str | None = Query(None, description="ค้นหาจากชื่ออำเภอ"),
    subdistrict_id: int | None = Query(None, description="กรองตามตำบล"),
    subdistrict_name: str | None = Query(None, description="ค้นหาจากชื่อตำบล"),
    subdistrict_postcode_id: int | None = Query(
        None, description="กรองตามแถว bridge sub_districts_postcode"
    ),
    postcode: str | None = Query(None, description="ค้นหาจากรหัสไปรษณีย์"),
    page: int = Query(1, ge=1),
    page_size: int | None = Query(None, ge=1),
    session: AsyncSession = Depends(get_session),
) -> DashboardCasesRead:
    """รายการคำร้องแบบ case_for_staff — ไม่ส่ง province_id = ทั้ง 77 จังหวัด มี pagination."""
    province_ids = await _validate_province_ids(session, province_id)
    resolved_page_size = min(page_size or settings.default_page_size, settings.max_page_size)
    filters = DashboardCaseFilters(
        province_ids=province_ids,
        current_status_ids=_clean_ids(current_status_id),
        type_money_ids=_clean_ids(type_money_id),
        case_number=_clean_text_filter(case_number),
        current_status=_clean_text_filter(current_status),
        firstname=_clean_text_filter(firstname),
        lastname=_clean_text_filter(lastname),
        cid=_clean_text_filter(cid),
        datetime_create=datetime_create,
        province_name=_clean_text_filter(province_name),
        district_id=district_id,
        district_name=_clean_text_filter(district_name),
        subdistrict_id=subdistrict_id,
        subdistrict_name=_clean_text_filter(subdistrict_name),
        subdistrict_postcode_id=subdistrict_postcode_id,
        postcode=_clean_text_filter(postcode),
    )
    total_items = await fetch_dashboard_cases_count(session, filters)
    total_pages = max(1, (total_items + resolved_page_size - 1) // resolved_page_size) if total_items else 1
    rows = await fetch_dashboard_cases_page(
        session,
        filters,
        limit=resolved_page_size,
        offset=(page - 1) * resolved_page_size,
    )
    return DashboardCasesRead(
        province_ids=province_ids,
        page=page,
        page_size=resolved_page_size,
        total_items=total_items,
        total_pages=total_pages,
        items=[_row_to_dashboard_case(row) for row in rows],
    )
